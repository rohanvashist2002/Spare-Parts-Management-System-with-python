import openpyxl

# Create a new Excel sheet for spare part data
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Spare Parts Data Sheet"

# Set the header row
sheet['A1'] = 'ID'
sheet['B1'] = 'Name'
sheet['C1'] = 'Description'                                                                                                         
sheet['D1'] = 'Unit Price'
sheet['E1'] = 'Quantity'
sheet['F1'] = 'Supplier'

# Create a dictionary to store spare part data
spare_parts = {}

def add_spare_part(id, name, description, unit_price, quantity,supplier):
    spare_parts[id] = {
        'name': name,
        'description': description,         
        'unit_price': unit_price,
        'quantity': quantity,
        'supplier': supplier
    }
    sheet.append([id, name, description, unit_price, quantity, supplier])
    wb.save('spare_parts.xlsx')

def update_spare_part(id, **kwargs):
    if id in spare_parts:
        for key, value in kwargs.items():
            spare_parts[id][key] = value
        row = sheet.max_row + 1
        sheet[row][0].value = id
        sheet[row][1].value = spare_parts[id]['name']
        sheet[row][2].value = spare_parts[id]['description']
        sheet[row][3].value = spare_parts[id]['unit_price']
        sheet[row][4].value = spare_parts[id]['quantity']
        sheet[row][5].value = spare_parts[id]['supplier']
        wb.save('spare_parts.xlsx')

def delete_spare_part(id):
    if id in spare_parts:
        del spare_parts[id]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == id:
                sheet.delete_rows(row[0])
                wb.save('spare_parts.xlsx')
                break
            
def get_all_spare_parts():
    output = ""
    for id, part in spare_parts.items():
        output += f"ID: {id}\n"
        output += f"Name: {part['name']}\n"
        output += f"Description: {part['description']}\n"
        output += f"Unit Price: {part['unit_price']}\n"
        output += f"Quantity: {part['quantity']}\n"
        output += f"Supplier: {part['supplier']}\n\n"
    return output

def get_spare_part(id):
    if id in spare_parts:
        part = spare_parts[id]
        output = f"ID: {id}\n"
        output += f"Name: {part['name']}\n"
        output += f"Description: {part['description']}\n"
        output += f"Unit Price: {part['unit_price']}\n"
        output += f"Quantity: {part['quantity']}\n"
        output += f"Supplier: {part['supplier']}\n"
        return output
    else:
        return "Spare part not found"


def main():
    while True:
        print("\nSpare Parts Management System")
        print("-------------------------------")
        print("1. Add a new spare part")
        print("2. Update an existing spare part")
        print("3. Delete a spare part")
        print("4. Get all spare parts")
        print("5. Get a specific spare part")
        print("6. Exit")

        choice = input("Enter your choice: ")

        if choice == "1":
            id = int(input("Enter the ID of the spare part: "))
            name = input("Enter the name of the spare part: ")
            description = input("Enter the description of the spare part: ")
            unit_price = float(input("Enter the unit price of the spare part: "))
            quantity = int(input("Enter the quantity of the spare part: "))
            supplier = input("Enter the supplier of the spare part: ")
            add_spare_part(id, name, description, unit_price, quantity, supplier)
        elif choice == "2":
            id = int(input("Enter the ID of the spare part to update: "))
            print("Enter the new values for the spare part (leave blank to keep current value):")
            name = input("Name: ")
            description = input("Description: ")
            unit_price = input("Unit Price: ")
            quantity = input("Quantity: ")
            supplier = input("Supplier: ")
            update_spare_part(id, name=name or None, description=description or None, unit_price=unit_price or None, quantity=quantity or None, supplier=supplier or None)
        elif choice == "3":
            id = int(input("Enter the ID of the spare part to delete: "))
            delete_spare_part(id)
        elif choice == "4":
            print(get_all_spare_parts())
        elif choice == "5":
            id = int(input("Enter the ID of the spare part to retrieve: "))
            print(get_spare_part(id))
        elif choice == "6":
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()